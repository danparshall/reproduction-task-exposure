"""Occupation profile loading and formatting for CDR classification prompts."""

from __future__ import annotations

import csv
from collections import defaultdict
from typing import Any


def load_occupation_profiles(csv_path: str) -> dict[str, dict[str, Any]]:
    """Load occupation profiles from CSV, keyed by O*NET SOC code.

    Args:
        csv_path: Path to occupation_profiles.csv

    Returns:
        Dict mapping SOC codes to profile dicts with fields:
        occupation_title, job_zone, top_gwas, physical_context,
        is_apprenticeable (bool), rapids_term_hours (str),
        carollo_licensed (str), carollo_n_states (str), carollo_pct (str)
    """
    profiles = {}
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            soc_code = row["onet_soc_code"]
            profiles[soc_code] = {
                "onet_soc_code": soc_code,
                "occupation_title": row["occupation_title"],
                "job_zone": row["job_zone"],
                "top_gwas": row["top_gwas"],
                "physical_context": row["physical_context"],
                "is_apprenticeable": row["is_apprenticeable"].strip().lower() == "true",
                "rapids_term_hours": row["rapids_term_hours"],
                "carollo_licensed": row.get("carollo_licensed", ""),
                "carollo_n_states": row.get("carollo_n_states", ""),
                "carollo_pct": row.get("carollo_pct", ""),
            }
    return profiles


def load_onet_descriptions(xlsx_path: str) -> dict[str, str]:
    """Load O*NET occupation descriptions from Occupation Data.xlsx.

    Args:
        xlsx_path: Path to Occupation Data.xlsx

    Returns:
        Dict mapping SOC codes to description strings
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    soc_idx = headers.index("O*NET-SOC Code")
    desc_idx = headers.index("Description")
    descriptions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        soc_code = row[soc_idx]
        description = row[desc_idx]
        if soc_code and description:
            descriptions[soc_code] = description
    wb.close()
    return descriptions


def format_occupation_profile(
    profile: dict[str, Any],
    description: str | None = None,
) -> str:
    """Format an occupation profile into a prompt header block.

    Args:
        profile: Dict with occupation profile fields
        description: Optional O*NET occupation description

    Returns:
        Formatted text block for prepending to user prompt
    """
    lines = []
    lines.append(f"OCCUPATION PROFILE: {profile['occupation_title']}")
    lines.append(f"SOC Code: {profile['onet_soc_code']}")
    lines.append(f"Job Zone: {profile['job_zone']}/5")

    if description:
        lines.append(f"O*NET Description: {description}")

    if profile.get("top_gwas"):
        lines.append("Top Work Activities (O*NET, Importance scale 1-5):")
        for entry in profile["top_gwas"].split(";"):
            entry = entry.strip()
            if entry:
                lines.append(f"  - {entry}")

    if profile["is_apprenticeable"]:
        hours = profile.get("rapids_term_hours", "").strip()
        if hours:
            lines.append(f"Registered Apprenticeship: Yes (term: {hours} hours)")
        else:
            lines.append("Registered Apprenticeship: Yes")
    else:
        lines.append("Registered Apprenticeship: No")

    # Carollo licensing status
    carollo_r = profile.get("carollo_licensed", "").strip()
    if carollo_r:
        n_states = profile.get("carollo_n_states", "").strip()
        if carollo_r == "R3":
            if n_states:
                n = int(n_states)
                pct_val = n / 51 * 100
                if pct_val >= 95:
                    lines.append(f"Regulatory Status: Licensed in {n}/51 states (universally regulated)")
                elif pct_val >= 50:
                    lines.append(f"Regulatory Status: Licensed in {n}/51 states ({pct_val:.0f}% of states)")
                else:
                    lines.append(f"Regulatory Status: Licensed in {n}/51 states (minority of states)")
            else:
                lines.append("Regulatory Status: Licensed (Carollo R3)")
        elif carollo_r == "R0_or_R1":
            lines.append("Regulatory Status: Not licensed (no state-level regulation)")

    return "\n".join(lines)


def group_tasks_by_occupation(tasks: list[dict]) -> dict[str, list[dict]]:
    """Group a flat list of tasks by their O*NET SOC code.

    Args:
        tasks: List of task dicts, each with 'onet_soc_code' key

    Returns:
        Dict mapping SOC codes to lists of task dicts
    """
    grouped: dict[str, list[dict]] = defaultdict(list)
    for task in tasks:
        grouped[task["onet_soc_code"]].append(task)
    return dict(grouped)
