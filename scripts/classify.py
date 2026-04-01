#!/usr/bin/env python3
"""CDR Classification Pipeline — Reproduce task exposure measurements.

Classifies O*NET occupational tasks on three axes:
  C (Cognitive complexity), D (Deployment difficulty), R (Regulatory restrictions)

Uses 3-model consensus (one model per provider: Anthropic, OpenAI, Google).
Supports checkpoint/resume, dry-run prompt inspection, and two-round consensus.

Usage:
  # Quick test — 12 occupations, ~$10 in API costs:
  python scripts/classify.py --soc-set expanded

  # Dry run — inspect prompts without API calls:
  python scripts/classify.py --dry-run

  # Full reproduction — all 923 occupations, ~$100 in API costs:
  python scripts/classify.py --soc-set full

  # Resume after interruption (auto-skips completed SOCs):
  python scripts/classify.py --soc-set full

  # Run single model (for partial resume):
  python scripts/classify.py --soc-set full --model gemini

  # Round 2 consensus on disputed tasks:
  python scripts/classify.py --soc-set full --round 2

  # 3-occupation pilot for initial API key validation:
  python scripts/classify.py --soc-set pilot
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from task_exposure.prompts import (  # noqa: E402
    expand_task_dwa_rows,
    format_cdr_system_prompt,
    format_user_prompt,
    get_response_format,
)
from task_exposure.clients import create_client  # noqa: E402
from task_exposure.profiles import (  # noqa: E402
    format_occupation_profile,
    load_occupation_profiles,
    load_onet_descriptions,
)
from task_exposure.runner import (  # noqa: E402
    build_consensus,
    build_prompt_snapshot,
    get_remaining_work,
    load_all_socs,
    load_checkpoints,
    save_aggregated_results,
    save_checkpoint,
)
from task_exposure.parser import parse_response  # noqa: E402

# ── Config ───────────────────────────────────────────────────────────────

# Load .env from project root (.env.local takes precedence)
for _env_name in (".env.local", ".env"):
    _env_path = PROJECT_ROOT / _env_name
    if _env_path.exists():
        load_dotenv(_env_path)
        break
else:
    print("NOTE: No .env file found. Copy .env.example to .env and add your API keys.")
    print(f"  cp {PROJECT_ROOT / '.env.example'} {PROJECT_ROOT / '.env'}")

ONET_DB = PROJECT_ROOT / "data" / "onet"
PROFILES_CSV = PROJECT_ROOT / "data" / "extracted" / "occupation_profiles.csv"
TASKS_CSV = PROJECT_ROOT / "data" / "extracted" / "onet_tasks_full_18k.csv"

DEFAULT_OUTPUT_DIR = "output/results"
DEFAULT_CHECKPOINT_DIR = "data/checkpoints"
MAX_CONCURRENT = 30  # Per provider — keeps under TPM/RPM limits

# ── SOC sets ─────────────────────────────────────────────────────────────
# "pilot" (3 occupations) — for API key validation, ~$2
# "expanded" (12 occupations) — for quick reproduction, ~$10
# "full" (923 occupations) — for complete reproduction, ~$100

SOC_SETS = {
    "pilot": [
        "39-5012.00",  # Hairdressers, Hairstylists, and Cosmetologists
        "49-9021.00",  # HVAC Mechanics and Installers
        "29-1141.00",  # Registered Nurses
    ],
    "expanded": [
        "39-5012.00",  # Hairdressers, Hairstylists, and Cosmetologists
        "49-9021.00",  # HVAC Mechanics and Installers
        "29-1141.00",  # Registered Nurses
        "11-1011.00",  # Chief Executives
        "13-2011.00",  # Accountants and Auditors
        "15-1252.00",  # Software Developers
        "25-2021.00",  # Elementary School Teachers
        "35-2014.00",  # Cooks, Restaurant
        "43-4051.00",  # Customer Service Representatives
        "47-2061.00",  # Construction Laborers
        "51-4121.00",  # Welders, Cutters, Solderers, and Brazers
        "53-3032.00",  # Heavy and Tractor-Trailer Truck Drivers
    ],
    "full": None,  # All SOCs from the CSV
}

# ── Model tiers ──────────────────────────────────────────────────────────
# Mid-tier: cheaper, good for reproduction and development (~$100 for full run)
# Flagship: expensive, for final validation (~$500+ for full run)

MODEL_TIERS = {
    "mid": {
        "sonnet": ("claude-sonnet-4-6", os.environ.get("ANTHROPIC_API_KEY", "")),
        "gpt": ("gpt-5-mini", os.environ.get("OPENAI_API_KEY", "")),
        "gemini": ("gemini-3-flash-preview", os.environ.get("GEMINI_API_KEY", "")),
    },
    "flagship": {
        "opus": ("claude-opus-4-6", os.environ.get("ANTHROPIC_API_KEY", "")),
        "gpt52": ("gpt-5.2", os.environ.get("OPENAI_API_KEY", "")),
        "gemini_pro": ("gemini-3-pro-preview", os.environ.get("GEMINI_API_KEY", "")),
    },
    "early2024": {
        "sonnet": ("claude-3-sonnet-20240229", os.environ.get("ANTHROPIC_API_KEY", "")),
        "gpt": ("gpt-3.5-turbo-0125", os.environ.get("OPENAI_API_KEY", "")),
        "gemini": ("gemini-1.0-pro", os.environ.get("GEMINI_API_KEY", "")),
    },
    "oldest": {
        "sonnet": ("claude-sonnet-4-20250514", os.environ.get("ANTHROPIC_API_KEY", "")),
        "gpt": ("gpt-4o-mini", os.environ.get("OPENAI_API_KEY", "")),
        "gemini": ("gemini-2.5-flash", os.environ.get("GEMINI_API_KEY", "")),
    },
    "oldest_o1": {
        "sonnet": ("claude-sonnet-4-20250514", os.environ.get("ANTHROPIC_API_KEY", "")),
        "gpt": ("o1", os.environ.get("OPENAI_API_KEY", "")),
        "gemini": ("gemini-2.5-flash", os.environ.get("GEMINI_API_KEY", "")),
    },
}

AXES = ["C", "D", "R"]


# ── O*NET data loaders ──────────────────────────────────────────────────


def load_task_to_dwas(xlsx_path: Path) -> dict[tuple[str, str], list[str]]:
    """Load (SOC, task_id) -> list of DWA Titles from Tasks to DWAs.xlsx."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    soc_idx = headers.index("O*NET-SOC Code")
    tid_idx = headers.index("Task ID")
    dwa_title_idx = headers.index("DWA Title")

    mapping: dict[tuple[str, str], list[str]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        soc = row[soc_idx]
        tid = str(int(row[tid_idx])) if isinstance(row[tid_idx], (int, float)) else str(row[tid_idx])
        dwa_title = row[dwa_title_idx]
        if dwa_title:
            mapping[(soc, tid)].append(dwa_title)
    wb.close()
    return dict(mapping)


def load_abilities(xlsx_path: Path, soc_codes: list[str], top_n: int = 10) -> dict[str, list[tuple[str, float]]]:
    """Load top-N abilities (Importance scale) per SOC code."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    soc_idx = headers.index("O*NET-SOC Code")
    name_idx = headers.index("Element Name")
    scale_idx = headers.index("Scale ID")
    val_idx = headers.index("Data Value")

    raw: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        soc = row[soc_idx]
        if soc in soc_codes and row[scale_idx] == "IM":
            raw[soc].append((row[name_idx], float(row[val_idx])))
    wb.close()

    return {soc: sorted(items, key=lambda x: x[1], reverse=True)[:top_n] for soc, items in raw.items()}


def load_work_context(xlsx_path: Path, soc_codes: list[str], top_n: int = 10) -> dict[str, list[tuple[str, float]]]:
    """Load top-N work context items (CX scale) per SOC code."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    soc_idx = headers.index("O*NET-SOC Code")
    name_idx = headers.index("Element Name")
    scale_idx = headers.index("Scale ID")
    val_idx = headers.index("Data Value")

    raw: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        soc = row[soc_idx]
        if soc in soc_codes and row[scale_idx] == "CX":
            raw[soc].append((row[name_idx], float(row[val_idx])))
    wb.close()

    return {soc: sorted(items, key=lambda x: x[1], reverse=True)[:top_n] for soc, items in raw.items()}


def annotate_tasks_with_dwas(
    grouped_tasks: dict[str, list[dict]],
    task_to_dwas: dict[tuple[str, str], list[str]],
) -> None:
    """In-place: add dwa_labels list to each task dict."""
    n_annotated = 0
    n_total = 0
    for soc, tasks in grouped_tasks.items():
        for task in tasks:
            n_total += 1
            key = (soc, task["task_id"])
            dwas = task_to_dwas.get(key, [])
            task["dwa_labels"] = dwas
            if dwas:
                n_annotated += 1
    print(f"  {n_annotated}/{n_total} tasks annotated with DWA labels")


# ── Model runner ────────────────────────────────────────────────────────


async def classify_occupation(
    client,
    model_label: str,
    soc: str,
    title: str,
    user_prompt: str,
    system_prompt: str,
    checkpoint_dir: str,
    reasoning_format: str = "per_axis",
) -> dict | None:
    """Classify one occupation, save checkpoint on success."""
    try:
        raw = await client.classify(user_prompt, system_prompt=system_prompt)
        if raw is None:
            print(f"  [{model_label}] {title}: EMPTY RESPONSE (model returned None)")
            return None
        parsed = parse_response(raw, reasoning_format=reasoning_format, axes=AXES)

        if parsed:
            save_checkpoint(parsed, soc, model_label, checkpoint_dir)
            print(f"  [{model_label}] {title}: {len(parsed)} rows classified")
        else:
            print(f"  [{model_label}] {title}: PARSE FAILED (empty result)")
        return parsed
    except Exception as e:
        print(f"  [{model_label}] {title} FAILED: {type(e).__name__}: {e!r}")
        return None


async def run_model(
    model_label: str,
    model_name: str,
    api_key: str,
    system_prompt: str,
    occupation_prompts: list[tuple[str, str, str]],  # (soc, title, user_prompt)
    checkpoint_dir: str,
    temperature: float | None = None,
    max_tokens: int = 16384,
    reasoning_format: str = "per_axis",
) -> dict[str, dict]:
    """Run all occupations through one model with concurrency limit."""
    client = create_client(model_name, api_key, timeout=300.0, temperature=temperature, max_tokens=max_tokens)
    sem = asyncio.Semaphore(MAX_CONCURRENT)
    all_results = {}

    async def run_one(soc, title, user_prompt):
        async with sem:
            result = await classify_occupation(
                client, model_label, soc, title, user_prompt,
                system_prompt, checkpoint_dir,
                reasoning_format=reasoning_format,
            )
            if result:
                all_results.update(result)

    await asyncio.gather(*(run_one(s, t, p) for s, t, p in occupation_prompts))
    await client.close()
    return all_results


# ── Round 2 consensus ──────────────────────────────────────────────────


def build_r2_prompt(
    soc: str,
    title: str,
    tasks: list[dict],
    model_results: dict[str, dict[str, dict[str, str]]],
    profile_text: str,
    abilities_text: str,
    work_context_text: str,
) -> str | None:
    """Build Round 2 (c-round) axis-dispute prompt for one occupation.

    Sends ALL non-unanimous task-axis combinations for re-rating. Each model
    sees the other models' ratings and reasoning (anonymized) for disputed
    axes only.

    Returns None if all task-axis combos in this occupation are unanimous.
    """
    rows = expand_task_dwa_rows(tasks)
    row_info = {row["row_id"]: (row["dwa_title"], row["task_description"]) for row in rows}

    all_task_ids = sorted(set().union(*(r.keys() for r in model_results.values())))
    soc_task_ids = [tid for tid in all_task_ids if tid in row_info]

    model_labels = list(model_results.keys())
    anon_labels = {label: f"Model {chr(65 + i)}" for i, label in enumerate(model_labels)}

    fully_agreed = []
    tasks_with_disputes = []

    for tid in soc_task_ids:
        ratings_by_model = {
            label: model_results[label][tid]
            for label in model_labels
            if tid in model_results[label]
        }
        if len(ratings_by_model) < 2:
            continue

        agreed_axes = {}
        disputed_axes = {}
        for axis in AXES:
            vals = {label: r.get(axis) for label, r in ratings_by_model.items() if r.get(axis)}
            unique = set(vals.values())
            if len(unique) == 1:
                agreed_axes[axis] = next(iter(unique))
            else:
                disputed_axes[axis] = {
                    label: {"rating": r[axis], "reasoning": r.get(f"{axis.lower()}_reasoning", "")}
                    for label, r in ratings_by_model.items()
                    if r.get(axis)
                }

        if not disputed_axes:
            first = next(iter(ratings_by_model.values()))
            fully_agreed.append((tid, first))
        else:
            tasks_with_disputes.append((tid, agreed_axes, disputed_axes))

    if not tasks_with_disputes:
        return None

    total_disputes = sum(len(d) for _, _, d in tasks_with_disputes)

    parts = [profile_text, ""]
    if abilities_text:
        parts += [abilities_text, ""]
    if work_context_text:
        parts += [work_context_text, ""]

    # Agreed tasks as reference
    parts.append("=== AGREED TASKS (reference only, do not re-rate) ===")
    parts.append("")
    parts.append("task_id | dwa_title | task_description | " + " | ".join(AXES))
    for tid, ratings in fully_agreed:
        dwa, desc = row_info.get(tid, ("", ""))
        axis_vals = " | ".join(ratings.get(a, "?") for a in AXES)
        parts.append(f"{tid} | {dwa} | {desc} | {axis_vals}")
    parts.append("")

    # Disputed tasks with per-model reasoning
    parts.append(f"=== DISPUTED AXES ({total_disputes} axis disputes across {len(tasks_with_disputes)} tasks) ===")
    parts.append("")
    for tid, agreed, disputed in tasks_with_disputes:
        dwa, desc = row_info.get(tid, ("", ""))
        parts.append(f"{tid} | {dwa} | {desc}")
        for axis in AXES:
            if axis in agreed:
                parts.append(f"  {axis}: AGREED = {agreed[axis]}")
            elif axis in disputed:
                parts.append(f"  {axis}: DISPUTED")
                for label in model_labels:
                    if label in disputed[axis]:
                        r = disputed[axis][label]
                        parts.append(f"    {anon_labels[label]}: {r['rating']} -- {r['reasoning']}")
        parts.append("")

    parts.append(
        f"Re-rate ONLY the {total_disputes} disputed axes above. "
        f"Consider each rater's reasoning, then provide your own independent "
        f"classification using the axis_dispute response format."
    )
    parts.append("")
    parts.append(get_response_format("axis_dispute", axes=AXES))

    return "\n".join(parts)


# ── Main ────────────────────────────────────────────────────────────────


async def main():
    parser = argparse.ArgumentParser(
        description="CDR Task Exposure Classification Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/classify.py --soc-set pilot      # 3 occupations, ~$2
  python scripts/classify.py --soc-set expanded    # 12 occupations, ~$10
  python scripts/classify.py --soc-set full        # 923 occupations, ~$100
  python scripts/classify.py --dry-run             # Inspect prompts, no API calls
  python scripts/classify.py --round 2             # Consensus on disputes
""",
    )
    parser.add_argument(
        "--soc-set",
        choices=list(SOC_SETS.keys()),
        default="expanded",
        help="SOC set: pilot (3, ~$2), expanded (12, ~$10), full (923, ~$100). Default: expanded.",
    )
    parser.add_argument(
        "--round",
        type=int,
        choices=[1, 2],
        default=1,
        help="Pipeline round: 1 (initial) or 2 (consensus on disputes).",
    )
    parser.add_argument(
        "--model-tier",
        choices=list(MODEL_TIERS.keys()),
        default="mid",
        help="Model tier: mid (sonnet/gpt-5-mini/flash) or flagship. Default: mid.",
    )
    parser.add_argument(
        "--model",
        type=str,
        default=None,
        help="Run single model (e.g. 'gemini', 'sonnet', 'gpt'). For partial resume.",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=None,
        help="Temperature override. Not recommended — see AGENTS.md 'Temperature' section. Default: provider defaults.",
    )
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=16384,
        help="Max output tokens per API call. Default: 16384.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Save prompts to output dir without calling APIs.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory for results. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--checkpoint-dir",
        type=str,
        default=DEFAULT_CHECKPOINT_DIR,
        help=f"Checkpoint directory. Default: {DEFAULT_CHECKPOINT_DIR}",
    )
    args = parser.parse_args()

    # Warn about temperature limitations:
    #   - Gemini: T=0 causes thinking-token exhaustion (create_client drops it)
    #   - Reasoning models (o-series, gpt-5-mini): API rejects T≠1 (client skips it)
    #   - Anthropic: T=0 is NOT deterministic in practice (~87% inter-run stability)
    # Temperature is not a useful lever for this pipeline. The --temperature flag
    # exists for experimentation but the defaults are the settled approach.
    if args.temperature is not None and args.temperature == 0.0:
        affected = []
        for name, _ in MODEL_TIERS[args.model_tier].values():
            if name.startswith("gemini"):
                affected.append(f"{name} (thinking-token exhaustion)")
            elif "mini" in name or name.startswith("o"):
                affected.append(f"{name} (reasoning model, only supports T=1)")
        if affected:
            print(
                "WARNING: --temperature 0 will be silently dropped for: "
                + ", ".join(affected)
            )
        print(
            "NOTE: T=0 does not meaningfully improve classification stability. "
            "See AGENTS.md 'Temperature' section for details."
        )

    models = MODEL_TIERS[args.model_tier]
    if args.model:
        if args.model not in models:
            valid = ", ".join(models.keys())
            print(f"ERROR: --model '{args.model}' not in tier '{args.model_tier}'. Valid: {valid}")
            sys.exit(1)
        models = {args.model: models[args.model]}

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    checkpoint_dir = args.checkpoint_dir

    # ── Load O*NET data ──────────────────────────────────────────────
    print("Loading O*NET data...")
    all_grouped = load_all_socs(str(TASKS_CSV))
    print(f"  {len(all_grouped)} total occupations, {sum(len(t) for t in all_grouped.values())} tasks")

    # Filter to selected SOC set
    soc_set = args.soc_set
    if SOC_SETS[soc_set] is not None:
        target_socs = SOC_SETS[soc_set]
        grouped_tasks = {s: all_grouped[s] for s in target_socs if s in all_grouped}
    else:
        grouped_tasks = all_grouped
        target_socs = sorted(grouped_tasks.keys())

    print(f"  Running {len(grouped_tasks)} occupations ({soc_set})")

    profiles = load_occupation_profiles(str(PROFILES_CSV))
    print(f"  {len(profiles)} occupation profiles")

    descriptions = load_onet_descriptions(str(ONET_DB / "Occupation Data.xlsx"))
    print(f"  {len(descriptions)} occupation descriptions")

    task_to_dwas = load_task_to_dwas(ONET_DB / "Tasks to DWAs.xlsx")
    print(f"  {len(task_to_dwas)} task->DWA mappings")
    annotate_tasks_with_dwas(grouped_tasks, task_to_dwas)

    soc_codes_set = set(target_socs)
    abilities = load_abilities(ONET_DB / "Abilities.xlsx", soc_codes_set)
    print(f"  Abilities loaded for {len(abilities)} occupations")

    work_context = load_work_context(ONET_DB / "Work Context.xlsx", soc_codes_set)
    print(f"  Work context loaded for {len(work_context)} occupations")

    # ── Build system prompt ──────────────────────────────────────────
    system_prompt = format_cdr_system_prompt(
        reasoning_format="per_axis",
        c_version="v2",
        r_version="v8",
    )
    print(f"\nSystem prompt (CDR, C-v2, R-v8): {len(system_prompt)} chars (~{len(system_prompt) // 4} tokens)")

    # Build one sample user prompt for snapshot
    sample_soc = target_socs[0]
    sample_tasks = grouped_tasks[sample_soc]
    sample_user_prompt = format_user_prompt(
        sample_tasks,
        axes=AXES,
        profile=profiles.get(sample_soc),
        description=descriptions.get(sample_soc),
        abilities=abilities.get(sample_soc),
        work_context=work_context.get(sample_soc),
    )

    # Save prompt snapshot
    snapshot = build_prompt_snapshot(system_prompt, sample_user_prompt)
    snapshot["timestamp"] = datetime.now(timezone.utc).isoformat()
    snapshot["prompt_versions"] = {"c": "v2", "d": "v4.1", "r": "v8"}
    snapshot["axes"] = AXES
    snapshot["reasoning_format"] = "per_axis"
    snapshot["soc_set"] = soc_set
    snapshot["n_occupations"] = len(grouped_tasks)
    snapshot["n_tasks"] = sum(len(t) for t in grouped_tasks.values())
    snapshot["models"] = {k: v[0] for k, v in models.items()}
    snapshot["temperature"] = args.temperature

    (out_dir / "prompt_snapshot.json").write_text(json.dumps(snapshot, indent=2))
    (out_dir / "system_prompt.txt").write_text(system_prompt)
    (out_dir / "sample_user_prompt.txt").write_text(sample_user_prompt)
    print(f"  Prompt snapshot saved to {out_dir}")
    print(f"  SHA-256: {snapshot['sha256'][:16]}...")

    if args.dry_run:
        print("\n=== DRY RUN -- prompts saved, no API calls ===")
        # Save all user prompts so they can be inspected
        prompts_dir = out_dir / "user_prompts"
        prompts_dir.mkdir(exist_ok=True)
        for soc in target_socs:
            tasks = grouped_tasks.get(soc, [])
            if not tasks:
                continue
            user_prompt = format_user_prompt(
                tasks,
                axes=AXES,
                profile=profiles.get(soc),
                description=descriptions.get(soc),
                abilities=abilities.get(soc),
                work_context=work_context.get(soc),
            )
            safe_name = f"{soc}_{tasks[0]['occupation_title'][:30].replace(' ', '_').replace(',', '')}"
            (prompts_dir / f"{safe_name}.txt").write_text(user_prompt)
        print(f"  {len(target_socs)} user prompts saved to {prompts_dir}")
        return

    # ── Round 1 ──────────────────────────────────────────────────────
    if args.round == 1:
        print(f"\n=== ROUND 1 -- {len(models)} models x {len(grouped_tasks)} occupations ===")

        async def run_one_model(label, model_info):
            model_name, api_key = model_info
            remaining = get_remaining_work(target_socs, checkpoint_dir, label)
            if not remaining:
                print(f"  [{label}] All {len(target_socs)} SOCs already checkpointed -- skipping")
                return
            print(f"  [{label}] {len(remaining)} SOCs remaining (of {len(target_socs)})")

            occupation_prompts = []
            for soc in remaining:
                tasks = grouped_tasks.get(soc, [])
                if not tasks:
                    continue
                user_prompt = format_user_prompt(
                    tasks,
                    axes=AXES,
                    profile=profiles.get(soc),
                    description=descriptions.get(soc),
                    abilities=abilities.get(soc),
                    work_context=work_context.get(soc),
                )
                title = tasks[0]["occupation_title"]
                occupation_prompts.append((soc, title, user_prompt))

            await run_model(label, model_name, api_key, system_prompt, occupation_prompts, checkpoint_dir, temperature=args.temperature, max_tokens=args.max_tokens)

        # Run all models in parallel (providers have independent rate limits)
        await asyncio.gather(*(run_one_model(label, info) for label, info in models.items()))

        # Build consensus from all checkpoints
        print("\n=== Building consensus ===")
        model_results = {}
        for label in MODEL_TIERS[args.model_tier]:
            checkpoints = load_checkpoints(checkpoint_dir, label)
            merged = {}
            for soc_results in checkpoints.values():
                merged.update(soc_results)
            if merged:
                model_results[label] = merged
                print(f"  [{label}] {len(merged)} total task rows")

        if len(model_results) >= 2:
            consensus, disputed = build_consensus(model_results, axes=AXES)
            save_aggregated_results(consensus, disputed, str(out_dir))
            total = len(consensus) + len(disputed)
            print(f"  Consensus: {len(consensus)}/{total} resolved ({100 * len(consensus) / total:.1f}%)")
            print(f"  Disputed: {len(disputed)}/{total} ({100 * len(disputed) / total:.1f}%)")
            print(f"  Results saved to {out_dir}")
        else:
            print(f"  Only {len(model_results)} model(s) have results -- need at least 2 for consensus")

    # ── Round 2 ──────────────────────────────────────────────────────
    elif args.round == 2:
        print("\n=== ROUND 2 -- consensus on disputes ===")

        # Load R1 checkpoints
        model_results = {}
        for label in MODEL_TIERS[args.model_tier]:
            checkpoints = load_checkpoints(checkpoint_dir, label)
            merged = {}
            for soc_results in checkpoints.values():
                merged.update(soc_results)
            if merged:
                model_results[label] = merged

        if len(model_results) < 3:
            print(f"  ERROR: Need 3 models for R2, found {len(model_results)}")
            sys.exit(1)

        # Find occupations with any non-unanimous task-axis combos
        r2_checkpoint_dir = checkpoint_dir + "_r2"
        r2_prompts = []
        for soc in target_socs:
            tasks = grouped_tasks.get(soc, [])
            if not tasks:
                continue
            title = tasks[0]["occupation_title"]

            # Build per-SOC model results
            soc_model_results = {}
            for label, all_model in model_results.items():
                rows = expand_task_dwa_rows(tasks)
                row_ids = {r["row_id"] for r in rows}
                soc_results = {tid: r for tid, r in all_model.items() if tid in row_ids}
                if soc_results:
                    soc_model_results[label] = soc_results

            profile_text = format_occupation_profile(
                profiles.get(soc, {}), descriptions.get(soc, "")
            ) if profiles.get(soc) else ""

            abilities_text = ""
            if abilities.get(soc):
                lines = ["Key Abilities (O*NET, Importance scale 1-5):"]
                for name, val in abilities[soc]:
                    lines.append(f"  - {name} ({val:.1f})")
                abilities_text = "\n".join(lines)

            wc_text = ""
            if work_context.get(soc):
                lines = ["Work Context (O*NET, frequency/extent scale 1-5):"]
                for name, val in work_context[soc]:
                    lines.append(f"  - {name} ({val:.1f})")
                wc_text = "\n".join(lines)

            r2_prompt = build_r2_prompt(
                soc, title, tasks, soc_model_results,
                profile_text, abilities_text, wc_text,
            )
            if r2_prompt:
                r2_prompts.append((soc, title, r2_prompt))

        print(f"  {len(r2_prompts)} occupations have disputes for R2")

        if not r2_prompts:
            print("  No disputes to resolve!")
            return

        # Run c-round through all models in parallel
        async def run_c_round_model(label, model_info):
            model_name, api_key = model_info
            remaining = get_remaining_work(
                [s for s, _, _ in r2_prompts], r2_checkpoint_dir, label,
            )
            prompts_to_run = [(s, t, p) for s, t, p in r2_prompts if s in remaining]
            if not prompts_to_run:
                print(f"  [{label}] c-round already complete")
                return
            print(f"  [{label}] Running c-round on {len(prompts_to_run)} occupations")
            await run_model(label, model_name, api_key, system_prompt, prompts_to_run, r2_checkpoint_dir, temperature=args.temperature, max_tokens=args.max_tokens, reasoning_format="axis_dispute")

        await asyncio.gather(*(run_c_round_model(label, info) for label, info in models.items()))

        print("  c-round complete. Run consensus aggregation separately.")

    # ── Run metadata ─────────────────────────────────────────────────
    metadata = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "soc_set": soc_set,
        "n_occupations": len(grouped_tasks),
        "n_tasks": sum(len(t) for t in grouped_tasks.values()),
        "round": args.round,
        "model_tier": args.model_tier,
        "models": {k: v[0] for k, v in models.items()},
        "prompt_versions": {"c": "v2", "d": "v4.1", "r": "v8"},
        "axes": AXES,
        "reasoning_format": "per_axis",
        "prompt_sha256": snapshot["sha256"],
    }
    (out_dir / "run_metadata.json").write_text(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    asyncio.run(main())
